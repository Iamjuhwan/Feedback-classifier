"""
Step 2b — targeted re-check.

Takes the labeled_v1.csv (your first pass) and pulls out rows whose ONLY
evidence of being health-related is a short/ambiguous keyword (e.g. "lafiya",
"arun", "ciwo") that's prone to false-positive substring matches inside
unrelated names/slang. These are the rows most likely to be mislabeled as
health-related when they aren't.

Everything else (rows that matched on an unambiguous keyword like "covid",
"vaccine", "hospital", "asibiti") is left as-is from your first pass.
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent))
from health_lexicon import keywords_for

REPO_ROOT = Path(__file__).resolve().parent.parent
LABELED_PATH = REPO_ROOT / "data" / "processed" / "labeled_v1.csv"
OUT_PATH = REPO_ROOT / "data" / "processed" / "recheck_weak_matches.xlsx"

WEAK_KEYWORDS = {"lafiya", "arun", "ciwo", "iba", "ogwu", "ọgwụ", "oria", "ọrịa", "magani", "oogun"}
LABELS = ["urgent_complaint", "general_feedback", "praise", "not_relevant"]
FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def matched_keywords(text, lang):
    tl = text.lower()
    return [kw for kw in keywords_for(lang) if kw.lower() in tl]


def main():
    with open(LABELED_PATH, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    flagged = []
    for r in rows:
        kws = matched_keywords(r["text"], r["language"])
        strong = [k for k in kws if k.lower() not in WEAK_KEYWORDS]
        if not strong:
            flagged.append(r)

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info.column_dimensions["A"].width = 100
    lines = [
        "RE-CHECK: WEAK-KEYWORD-ONLY MATCHES",
        "",
        f"{len(flagged)} of your {len(rows)} first-pass labels only matched a short/ambiguous "
        "keyword (lafiya, arun, ciwo, iba, ogwu, oogun, magani...) that commonly false-positives "
        "inside names, slang, or unrelated words.",
        "",
        "For each row: check if it's ACTUALLY about health/community-service feedback.",
        "  - If yes, keep or fix the final_label as before.",
        "  - If no (money talk, gossip, politics, unrelated slang etc.) set final_label to "
        "'not_relevant' — don't just leave your old label in place.",
        "",
        "Everything NOT in this file (rows that matched a clear keyword like covid, vaccine, "
        "hospital, asibiti, dokita) is being kept from your first pass as-is.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws_info.cell(row=i, column=1, value=line)
        cell.font = Font(name=FONT_NAME, bold=(i == 1), size=14 if i == 1 else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Recheck")
    headers = ["row_id", "text", "language", "your_first_pass_label", "final_label"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL

    for i, r in enumerate(flagged, start=2):
        ws.cell(row=i, column=1, value=i - 1).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=2, value=r["text"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=3, value=r["language"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=4, value=r["final_label"]).font = Font(name=FONT_NAME)
        fc = ws.cell(row=i, column=5, value=r["final_label"])  # pre-filled with old value, editable
        fc.font = Font(name=FONT_NAME)
        fc.fill = INPUT_FILL

    widths = {"A": 8, "B": 85, "C": 10, "D": 20, "E": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    dv = DataValidation(type="list", formula1=f'"{",".join(LABELS)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E2:E{len(flagged)+1}")

    wb.save(OUT_PATH)
    print(f"Wrote {len(flagged)} rows to re-check: {OUT_PATH}")


if __name__ == "__main__":
    main()
