"""
Step 2 of the pipeline.

Turns data/processed/health_feedback_candidates.csv into an Excel workbook
for manual relabeling: dropdown-validated final_label column, instructions
tab, a random sample-order so review isn't biased by language grouping.
"""

import csv
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
IN_PATH = REPO_ROOT / "data" / "processed" / "health_feedback_candidates.csv"
OUT_PATH = REPO_ROOT / "data" / "processed" / "relabeling_review.xlsx"

LABELS = ["urgent_complaint", "general_feedback", "praise", "not_relevant"]

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def main():
    with open(IN_PATH, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    random.seed(42)
    random.shuffle(rows)  # de-bias language-grouped review order

    wb = Workbook()

    # --- Instructions sheet ---
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info.column_dimensions["A"].width = 100
    lines = [
        "COMMUNITY HEALTH FEEDBACK — MANUAL RELABELING",
        "",
        "Source: NaijaSenti (real, human-annotated Nigerian-language tweets), filtered to a "
        "health-domain subset by keyword matching.",
        "",
        "The 'suggested_label' column is a heuristic guess (negative sentiment -> "
        "urgent_complaint, neutral -> general_feedback, positive -> praise). It WILL be wrong "
        "often — sentiment isn't the same thing as urgency, and the keyword filter lets some "
        "non-health text through.",
        "",
        "For each row on the 'Review' tab:",
        "  1. Read the text.",
        "  2. Pick the correct value in 'final_label' from the dropdown:",
        "       - urgent_complaint : reports a health problem needing fast attention",
        "         (outbreak, drug shortage, facility failure, serious symptom)",
        "       - general_feedback : health-related but not urgent (opinion, question, routine remark)",
        "       - praise           : positive feedback about a health service/campaign/outcome",
        "       - not_relevant     : the keyword filter false-positived — not actually about health",
        "  3. Leave 'final_label' blank to skip a row for now.",
        "",
        "Aim for at least 300-500 labeled rows for a usable v1 training set; label more over time.",
        "Rows are pre-shuffled so you're not reviewing one language in a solid block.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws_info.cell(row=i, column=1, value=line)
        cell.font = Font(name=FONT_NAME, bold=(i == 1), size=14 if i == 1 else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Review sheet ---
    ws = wb.create_sheet("Review")
    headers = ["row_id", "text", "language", "source_sentiment", "suggested_label", "final_label"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=i - 1).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=2, value=row["text"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=3, value=row["language"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=4, value=row["source_sentiment"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=5, value=row["suggested_label"]).font = Font(name=FONT_NAME)
        final_cell = ws.cell(row=i, column=6, value="")
        final_cell.font = Font(name=FONT_NAME)
        final_cell.fill = INPUT_FILL

    # column widths
    widths = {"A": 8, "B": 80, "C": 10, "D": 16, "E": 18, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # dropdown validation on final_label column
    dv = DataValidation(type="list", formula1=f'"{",".join(LABELS)}"', allow_blank=True)
    ws.add_data_validation(dv)
    last_row = len(rows) + 1
    dv.add(f"F2:F{last_row}")

    wb.save(OUT_PATH)
    print(f"Wrote {len(rows)} rows to review: {OUT_PATH}")


if __name__ == "__main__":
    main()
